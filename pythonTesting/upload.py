import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


def updateDatafromExcel(filePath, colName, searchTerm, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}
    for i in range(1, sheet.max_column+ 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i
    for i in range(1, sheet.max_row+ 1):
        for j in range(1, sheet.max_column+ 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(Dict["row"], Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:/Users/abdul.saboor/Downloads/download.xlsx"
fruit_name = 'Apple'
new_price = "999"
driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()
updateDatafromExcel(file_path, "price", fruit_name, new_price)
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
WebDriverWait(driver, 5).until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/following-sibling::div[2]/div").text
assert actual_price == new_price
