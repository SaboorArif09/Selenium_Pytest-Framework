import openpyxl


class HomePageData:

    test_homePage_data = [{"name":"saboor", "email":"heloo@gmail.com", "password":"12345", "gender":"Male"},{"name":"abdul", "email":"hi@gmail.com", "password":"12345", "gender":"Female"}]

    @staticmethod
    def getDatafromExcel(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\abdul.saboor\\Downloads\\pythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]