import openpyxl

book = openpyxl.load_workbook("C:\\Users\\abdul.saboor\\Downloads\\pythonDemo.xlsx")
sheet = book.active
Dict={}
read = sheet.cell(row=1, column=2)
print(read.value)

write = sheet.cell(row=2, column=2).value = "saboor"
print(write)

print(sheet['A3'].value)

print(sheet.max_column)
print(sheet.max_row)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value

print(Dict)